from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO


class XLSXWriter(object):
    def __init__(self, fieldnames=[]):
        self.wb = Workbook()
        self.sheet = self.wb.worksheets[0]
        self.current_row = 1
        self.fieldnames = fieldnames

    def writerow(self, row):
        for i, field in enumerate(self.fieldnames, 1):
            cell = self.sheet.cell(column=i, row=self.current_row)
            cell.value = row.get(field)
            col_key = get_column_letter(cell.column)
            self.sheet.column_dimensions[col_key].width = max(self.sheet.column_dimensions[col_key].width,
                                                              len(str(cell.value)) + 4)
        self.current_row += 1

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)

    def get_file(self):
        file = BytesIO()
        self.wb.save(file)
        return file
